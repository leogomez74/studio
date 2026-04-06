from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment

wb = Workbook()

# ─── ESTILOS ─────────────────────────────────────────────────────────────────
header_fill   = PatternFill('solid', start_color='1e3a5f')
required_fill = PatternFill('solid', start_color='fee2e2')
header_font   = Font(name='Arial', bold=True, color='FFFFFF', size=10)
req_font      = Font(name='Arial', bold=True, color='b91c1c', size=10)
data_font     = Font(name='Arial', size=10)
center        = Alignment(horizontal='center', vertical='center', wrap_text=True)
left          = Alignment(horizontal='left', vertical='center')
thin          = Side(style='thin', color='CBD5E1')
border        = Border(left=thin, right=thin, top=thin, bottom=thin)

# ─── HOJA 1: Plantilla ───────────────────────────────────────────────────────
ws = wb.active
ws.title = 'Plantilla'

# (key, label, requerido, ancho, ejemplo)
cols = [
    ('cedula',              'Cedula *',              True,  18, '1-0234-5678'),
    ('email',               'Email *',               True,  28, 'juan.perez@correo.com'),
    ('telefono_principal',  'Telefono Principal *',  True,  20, '8888-8888'),
    ('nombre',              'Nombre',                False, 18, 'Juan'),
    ('apellido1',           'Primer Apellido',       False, 18, 'Perez'),
    ('apellido2',           'Segundo Apellido',      False, 18, 'Rodriguez'),
    ('fecha_nacimiento',    'Fecha Nacimiento',      False, 18, '1990-05-20'),
    ('genero',              'Genero',                False, 14, 'Masculino'),
    ('estado_civil',        'Estado Civil',          False, 16, 'Casado'),
    ('nacionalidad',        'Nacionalidad',          False, 16, 'Costarricense'),
    ('cedula_vencimiento',  'Vencimiento Cedula',    False, 18, '2028-01-15'),
    ('whatsapp',            'WhatsApp',              False, 16, '8888-8888'),
    ('tel_casa',            'Tel. Casa',             False, 16, '2222-3333'),
    ('tel_amigo',           'Tel. Amigo/Ref.',       False, 18, '7777-6666'),
    ('telefono2',           'Telefono 2',            False, 16, '6666-5555'),
    ('telefono3',           'Telefono 3',            False, 16, '5555-4444'),
    ('provincia',           'Provincia',             False, 16, 'San Jose'),
    ('canton',              'Canton',                False, 16, 'Escazu'),
    ('distrito',            'Distrito',              False, 16, 'San Rafael'),
    ('direccion1',          'Direccion 1',           False, 32, '100m norte del parque'),
    ('direccion2',          'Direccion 2',           False, 32, 'Casa azul, porton negro'),
    ('ocupacion',           'Ocupacion',             False, 20, 'Docente'),
    ('institucion_labora',  'Institucion Laboral',   False, 28, 'MEP'),
    ('departamento_cargo',  'Departamento/Cargo',    False, 24, 'Direccion Regional'),
    ('puesto',              'Puesto',                False, 22, 'Maestro de Ensenanza'),
    ('estado_puesto',       'Estado Puesto',         False, 16, 'Propietario'),
    ('sector',              'Sector',                False, 14, 'Publico'),
    ('nivel_academico',     'Nivel Academico',       False, 18, 'Universidad'),
    ('profesion',           'Profesion',             False, 20, 'Educacion'),
    ('actividad_economica', 'Actividad Economica',   False, 22, 'Educacion publica'),
    ('trabajo_provincia',   'Prov. Trabajo',         False, 16, 'San Jose'),
    ('trabajo_canton',      'Canton Trabajo',        False, 16, 'Central'),
    ('trabajo_distrito',    'Distrito Trabajo',      False, 16, 'Carmen'),
    ('trabajo_direccion',   'Direccion Trabajo',     False, 30, 'Av. 2, Ministerio de Ed.'),
    ('source',              'Fuente',                False, 16, 'Referido'),
    ('notes',               'Notas',                 False, 32, 'Cliente referido por amigo'),
    ('relacionado_a',       'Referido por',          False, 22, 'Maria Gonzalez'),
    ('tipo_relacion',       'Tipo Relacion',         False, 18, 'Amiga'),
]

ws.row_dimensions[1].height = 36
ws.row_dimensions[2].height = 20
ws.freeze_panes = 'A2'

for col_idx, (key, label, required, width, example) in enumerate(cols, start=1):
    cell = ws.cell(row=1, column=col_idx, value=label)
    cell.fill      = required_fill if required else header_fill
    cell.font      = req_font if required else header_font
    cell.alignment = center
    cell.border    = border

    ex = ws.cell(row=2, column=col_idx, value=example)
    ex.font      = data_font
    ex.alignment = left
    ex.border    = border

    ws.column_dimensions[get_column_letter(col_idx)].width = width

    if required:
        comment = Comment('CAMPO OBLIGATORIO\nNo dejar en blanco.', 'CR Studio')
        comment.width  = 180
        comment.height = 50
        cell.comment = comment

# ─── HOJA 2: Instrucciones ───────────────────────────────────────────────────
wi = wb.create_sheet('Instrucciones')
wi.column_dimensions['A'].width = 85

title_f   = Font(name='Arial', bold=True, size=14, color='1e3a5f')
section_f = Font(name='Arial', bold=True, size=11, color='1e3a5f')
body_f    = Font(name='Arial', size=10)
red_f     = Font(name='Arial', bold=True, size=10, color='b91c1c')

wi['A1'] = 'Instrucciones para Importacion Masiva de Clientes - CR Studio'
wi['A1'].font = title_f
wi.row_dimensions[1].height = 28

wi['A3'] = 'REGLAS GENERALES'
wi['A3'].font = section_f

reglas = [
    '  Los campos marcados con * y fondo rojo en la plantilla son OBLIGATORIOS.',
    '  La cedula debe ser unica - no puede repetirse ni existir ya en el sistema.',
    '  El email debe ser unico - no puede repetirse ni existir ya en el sistema.',
    '  Las fechas deben estar en formato YYYY-MM-DD  (ejemplo: 1990-05-20).',
    '  No eliminar ni reordenar las columnas de la hoja Plantilla.',
    '  No modificar la fila 1 (encabezados).',
    '  Llenar los datos desde la fila 2 (la fila 2 es un ejemplo, puede reemplazarla).',
    '  Maximo recomendado: 500 registros por archivo.',
    '  Usar los valores exactos indicados en la hoja Valores_Validos para campos con lista.',
]
for i, r in enumerate(reglas, start=4):
    wi.cell(row=i, column=1, value=r).font = body_f
    wi.row_dimensions[i].height = 16

wi['A14'] = 'PROCESO DE IMPORTACION'
wi['A14'].font = section_f

pasos = [
    '  1. Complete la plantilla en la hoja Plantilla desde la fila 2.',
    '  2. Guarde el archivo en formato .xlsx (Excel).',
    '  3. En el sistema CR Studio, vaya al modulo de Clientes / CRM.',
    '  4. Busque la opcion Importar clientes y seleccione este archivo.',
    '  5. Revise la vista previa que muestra el sistema antes de confirmar.',
    '  6. Confirme la importacion. El sistema indicara los errores si los hay.',
]
for i, p in enumerate(pasos, start=15):
    wi.cell(row=i, column=1, value=p).font = body_f
    wi.row_dimensions[i].height = 16

wi['A22'] = 'CAMPOS OBLIGATORIOS'
wi['A22'].font = section_f
obligatorios = [
    '  cedula              -> Numero de identificacion (ej: 1-0234-5678)',
    '  email               -> Correo electronico valido',
    '  telefono_principal  -> Numero de telefono principal (ej: 8888-8888)',
]
for i, o in enumerate(obligatorios, start=23):
    wi.cell(row=i, column=1, value=o).font = red_f
    wi.row_dimensions[i].height = 16

# ─── HOJA 3: Valores_Validos ─────────────────────────────────────────────────
wv = wb.create_sheet('Valores_Validos')
wv.column_dimensions['A'].width = 26
wv.column_dimensions['B'].width = 65

h1 = wv.cell(row=1, column=1, value='Campo')
h2 = wv.cell(row=1, column=2, value='Valores Aceptados (escribir exactamente como aparece)')
for h in [h1, h2]:
    h.font      = Font(name='Arial', bold=True, size=10, color='FFFFFF')
    h.fill      = header_fill
    h.alignment = center
    h.border    = border
wv.row_dimensions[1].height = 22

alt_fill = PatternFill('solid', start_color='EFF6FF')
valores = [
    ('genero',            'Masculino | Femenino | Otro'),
    ('estado_civil',      'Soltero | Casado | Divorciado | Viudo | Union Libre'),
    ('estado_puesto',     'Propietario | Interino | Contratado'),
    ('sector',            'Publico | Privado | Independiente'),
    ('nivel_academico',   'Primaria | Secundaria | Tecnico | Universidad | Postgrado'),
    ('source',            'Referido | Web | Llamada | Visita | Redes Sociales | Otro'),
    ('fecha_nacimiento',  'Formato: YYYY-MM-DD  (ej: 1990-05-20)'),
    ('cedula_vencimiento','Formato: YYYY-MM-DD  (ej: 2028-01-15)'),
]
for i, (campo, vals) in enumerate(valores, start=2):
    c1 = wv.cell(row=i, column=1, value=campo)
    c2 = wv.cell(row=i, column=2, value=vals)
    for c in [c1, c2]:
        c.font      = Font(name='Arial', size=10)
        c.border    = border
        c.alignment = left
        if i % 2 == 0:
            c.fill = alt_fill
    wv.row_dimensions[i].height = 16

wb.save('public/plantilla_importacion_clientes.xlsx')
print('Archivo generado: public/plantilla_importacion_clientes.xlsx')
