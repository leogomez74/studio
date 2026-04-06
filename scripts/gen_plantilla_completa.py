"""
Genera plantilla_carga_masiva_completa.xlsx
Hoja 1: DATOS (una fila por caso completo: cliente + oportunidad + analisis + credito)
Hoja 2: Instrucciones
Hoja 3: Valores_Validos
Hoja 4: Ejemplo_Completo
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment

wb = Workbook()

# ─── ESTILOS ─────────────────────────────────────────────────────────────────
def fill(color): return PatternFill('solid', start_color=color)
def font(color='000000', bold=False, size=10): return Font(name='Arial', color=color, bold=bold, size=size)

center = Alignment(horizontal='center', vertical='center', wrap_text=True)
left   = Alignment(horizontal='left',   vertical='center', wrap_text=False)
thin   = Side(style='thin', color='CBD5E1')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# Colores por sección
COLOR_CLIENTE    = '1e3a5f'  # azul oscuro
COLOR_OPP        = '065f46'  # verde oscuro
COLOR_ANALISIS   = '7c3aed'  # violeta
COLOR_CREDITO    = '9a3412'  # naranja oscuro
COLOR_REQ        = 'fee2e2'  # rojo claro (campo obligatorio)
COLOR_OPP_BG     = 'd1fae5'  # verde claro
COLOR_ANA_BG     = 'ede9fe'  # violeta claro
COLOR_CRED_BG    = 'ffedd5'  # naranja claro

# ─── DEFINICIÓN DE COLUMNAS ───────────────────────────────────────────────────
# (key, label, requerido, ancho, seccion, ejemplo)
cols = [
    # ── CLIENTE ──────────────────────────────────────────────────────────────
    ('cedula',              'Cedula *',              True,  18, 'CLIENTE',   '1-0234-5678'),
    ('email',               'Email *',               True,  28, 'CLIENTE',   'juan.perez@correo.com'),
    ('telefono_principal',  'Telefono *',            True,  18, 'CLIENTE',   '8888-8888'),
    ('nombre',              'Nombre',                False, 16, 'CLIENTE',   'Juan'),
    ('apellido1',           'Apellido 1',            False, 16, 'CLIENTE',   'Perez'),
    ('apellido2',           'Apellido 2',            False, 16, 'CLIENTE',   'Rodriguez'),
    ('fecha_nacimiento',    'Fecha Nacimiento',      False, 18, 'CLIENTE',   '1990-05-20'),
    ('genero',              'Genero',                False, 14, 'CLIENTE',   'Masculino'),
    ('estado_civil',        'Estado Civil',          False, 16, 'CLIENTE',   'Casado'),
    ('nacionalidad',        'Nacionalidad',          False, 16, 'CLIENTE',   'Costarricense'),
    ('whatsapp',            'WhatsApp',              False, 16, 'CLIENTE',   '8888-8888'),
    ('tel_casa',            'Tel. Casa',             False, 16, 'CLIENTE',   '2222-3333'),
    ('provincia',           'Provincia',             False, 16, 'CLIENTE',   'San Jose'),
    ('canton',              'Canton',                False, 16, 'CLIENTE',   'Escazu'),
    ('distrito',            'Distrito',              False, 16, 'CLIENTE',   'San Rafael'),
    ('direccion1',          'Direccion',             False, 30, 'CLIENTE',   '100m norte del parque'),
    ('ocupacion',           'Ocupacion',             False, 18, 'CLIENTE',   'Docente'),
    ('institucion_labora',  'Institucion Laboral',   False, 24, 'CLIENTE',   'MEP'),
    ('puesto',              'Puesto',                False, 20, 'CLIENTE',   'Maestro'),
    ('estado_puesto',       'Estado Puesto',         False, 16, 'CLIENTE',   'Propietario'),
    ('sector',              'Sector',                False, 14, 'CLIENTE',   'Publico'),
    ('nivel_academico',     'Nivel Academico',       False, 18, 'CLIENTE',   'Universidad'),
    ('profesion',           'Profesion',             False, 18, 'CLIENTE',   'Educacion'),
    ('source',              'Fuente',                False, 14, 'CLIENTE',   'Referido'),

    # ── OPORTUNIDAD ──────────────────────────────────────────────────────────
    ('opp_amount',          'Monto Solicitado *',    True,  20, 'OPORTUNIDAD', '3500000'),
    ('opp_type',            'Tipo Oportunidad',      False, 22, 'OPORTUNIDAD', 'Credito Personal'),
    ('opp_vertical',        'Vertical/Institucion',  False, 22, 'OPORTUNIDAD', 'MEP'),
    ('opp_status',          'Estado Oportunidad',    False, 20, 'OPORTUNIDAD', 'Pendiente'),
    ('opp_close_date',      'Fecha Cierre Esperada', False, 20, 'OPORTUNIDAD', '2026-04-30'),
    ('opp_comments',        'Comentarios Opp.',      False, 28, 'OPORTUNIDAD', 'Cliente interesado en credito 36 meses'),

    # ── ANALISIS ─────────────────────────────────────────────────────────────
    ('ana_monto_solicitado','Monto Analisis *',      True,  20, 'ANALISIS',  '3500000'),
    ('ana_plazo',           'Plazo (meses) *',       True,  16, 'ANALISIS',  '36'),
    ('ana_cuota',           'Cuota Estimada',        False, 18, 'ANALISIS',  '125000'),
    ('ana_divisa',          'Divisa',                False, 12, 'ANALISIS',  'CRC'),
    ('ana_ingreso_bruto',   'Ingreso Bruto',         False, 18, 'ANALISIS',  '800000'),
    ('ana_ingreso_neto',    'Ingreso Neto',          False, 18, 'ANALISIS',  '620000'),
    ('ana_ingreso_bruto_2', 'Ingreso Bruto 2',       False, 18, 'ANALISIS',  ''),
    ('ana_ingreso_neto_2',  'Ingreso Neto 2',        False, 18, 'ANALISIS',  ''),
    ('ana_ingreso_bruto_3', 'Ingreso Bruto 3',       False, 18, 'ANALISIS',  ''),
    ('ana_ingreso_neto_3',  'Ingreso Neto 3',        False, 18, 'ANALISIS',  ''),
    ('ana_cargo',           'Cargo Credid',          False, 20, 'ANALISIS',  'Docente'),
    ('ana_nombramiento',    'Nombramiento',          False, 20, 'ANALISIS',  'Propietario'),
    ('ana_numero_manchas',  'Num. Manchas',          False, 14, 'ANALISIS',  '0'),
    ('ana_numero_juicios',  'Num. Juicios',          False, 14, 'ANALISIS',  '0'),
    ('ana_numero_embargos', 'Num. Embargos',         False, 14, 'ANALISIS',  '0'),
    ('ana_estado_pep',      'Estado PEP',            False, 16, 'ANALISIS',  'Pendiente'),
    ('ana_propuesta',       'Propuesta',             False, 28, 'ANALISIS',  'Se aprueba credito por buen perfil'),
    ('ana_description',     'Descripcion Analisis',  False, 28, 'ANALISIS',  'Solicitante estable con buena capacidad de pago'),

    # ── CREDITO ──────────────────────────────────────────────────────────────
    ('cred_monto',          'Monto Credito *',       True,  18, 'CREDITO',   '3500000'),
    ('cred_cuota',          'Cuota *',               True,  16, 'CREDITO',   '125000'),
    ('cred_plazo',          'Plazo (meses) *',       True,  16, 'CREDITO',   '36'),
    ('cred_tasa_anual',     'Tasa Anual % *',        True,  16, 'CREDITO',   '24'),
    ('cred_tipo',           'Tipo Credito',          False, 20, 'CREDITO',   'Personal'),
    ('cred_formalized_at',  'Fecha Formalizacion',   False, 20, 'CREDITO',   '2026-04-01'),
    ('cred_fecha_culminacion','Fecha Vencimiento',   False, 20, 'CREDITO',   '2029-04-01'),
    ('cred_deductora',      'Deductora (nombre)',    False, 22, 'CREDITO',   'COOPESANGABRIEL R.L.'),
    ('cred_poliza',         'Poliza (1=Si, 0=No)',   False, 18, 'CREDITO',   '1'),
    ('cred_garantia',       'Garantia',              False, 22, 'CREDITO',   'Fiduciaria'),
    ('cred_status',         'Estado Credito',        False, 18, 'CREDITO',   'Activo'),
    ('cred_description',    'Descripcion Credito',   False, 28, 'CREDITO',   'Credito personal aprobado y formalizado'),
]

SECTION_COLORS = {
    'CLIENTE':    (COLOR_CLIENTE, 'FFFFFF'),
    'OPORTUNIDAD':(COLOR_OPP,     'FFFFFF'),
    'ANALISIS':   (COLOR_ANALISIS,'FFFFFF'),
    'CREDITO':    (COLOR_CREDITO, 'FFFFFF'),
}
SECTION_BG = {
    'CLIENTE':    None,
    'OPORTUNIDAD': COLOR_OPP_BG,
    'ANALISIS':    COLOR_ANA_BG,
    'CREDITO':     COLOR_CRED_BG,
}

# ─── HOJA 1: DATOS ───────────────────────────────────────────────────────────
ws = wb.active
ws.title = 'DATOS'

# Fila 1: labels de sección (agrupados)
# Fila 2: encabezados de columna
# Fila 3 en adelante: datos

# Calcular rangos por sección
section_ranges = {}
for col_idx, (key, label, req, width, section, example) in enumerate(cols, start=1):
    if section not in section_ranges:
        section_ranges[section] = [col_idx, col_idx]
    else:
        section_ranges[section][1] = col_idx

# Fila 1 — banners de sección
for section, (start_col, end_col) in section_ranges.items():
    bg, fg = SECTION_COLORS[section]
    ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
    cell = ws.cell(row=1, column=start_col, value=f'--- {section} ---')
    cell.fill      = fill(bg)
    cell.font      = Font(name='Arial', bold=True, color=fg, size=11)
    cell.alignment = center
ws.row_dimensions[1].height = 22

# Fila 2 — encabezados de columna
ws.row_dimensions[2].height = 36
ws.freeze_panes = 'A3'

for col_idx, (key, label, req, width, section, example) in enumerate(cols, start=1):
    bg, fg = SECTION_COLORS[section]

    cell = ws.cell(row=2, column=col_idx, value=label)
    cell.fill      = fill(COLOR_REQ) if req else fill(bg)
    cell.font      = Font(name='Arial', bold=True, color=('b91c1c' if req else fg), size=9)
    cell.alignment = center
    cell.border    = border
    ws.column_dimensions[get_column_letter(col_idx)].width = width

    if req:
        comment = Comment('CAMPO OBLIGATORIO', 'CR Studio')
        comment.width  = 160
        comment.height = 40
        cell.comment = comment

# Fila 3 — fila de ejemplo (gris claro)
ex_fill = fill('F8FAFC')
for col_idx, (key, label, req, width, section, example) in enumerate(cols, start=1):
    c = ws.cell(row=3, column=col_idx, value=example)
    c.font      = Font(name='Arial', size=9, italic=True, color='64748b')
    c.fill      = ex_fill
    c.alignment = left
    c.border    = border
ws.row_dimensions[3].height = 16

# ─── HOJA 2: Instrucciones ───────────────────────────────────────────────────
wi = wb.create_sheet('Instrucciones')
wi.column_dimensions['A'].width = 90

wi['A1'] = 'Plantilla de Carga Masiva Completa — CR Studio'
wi['A1'].font = Font(name='Arial', bold=True, size=14, color='1e3a5f')
wi.row_dimensions[1].height = 28

wi['A3'] = 'QUE INCLUYE ESTE ARCHIVO'
wi['A3'].font = Font(name='Arial', bold=True, size=11, color='1e3a5f')
secciones = [
    '  CLIENTE      -> Datos personales, contacto, laboral y ubicacion del cliente/lead',
    '  OPORTUNIDAD  -> Monto solicitado, tipo, institucion y estado de la oportunidad',
    '  ANALISIS     -> Ingresos, manchas, propuesta crediticia y datos de analisis',
    '  CREDITO      -> Monto, cuota, plazo, tasa, deductora y condiciones del credito',
]
for i, s in enumerate(secciones, start=4):
    c = wi.cell(row=i, column=1, value=s)
    c.font = Font(name='Arial', size=10)
wi.row_dimensions[4].height = 14

wi['A9'] = 'REGLAS OBLIGATORIAS'
wi['A9'].font = Font(name='Arial', bold=True, size=11, color='1e3a5f')
reglas = [
    '  Los campos marcados con * y fondo rojo son OBLIGATORIOS — no dejarlos vacios.',
    '  La cedula y el email deben ser unicos — no pueden existir ya en el sistema.',
    '  Cada fila representa UN caso completo: 1 cliente + 1 oportunidad + 1 analisis + 1 credito.',
    '  Si un campo no aplica, dejarlo en blanco (no escribir N/A ni guiones).',
    '  Fechas en formato YYYY-MM-DD  (ejemplo: 2026-04-01).',
    '  Montos sin simbolos ni comas — solo el numero  (ejemplo: 3500000).',
    '  Tasa anual como porcentaje entero  (ejemplo: 24 para 24%).',
    '  Poliza: escribir 1 para Si, 0 para No.',
    '  Deductora: escribir el nombre exacto como aparece en el sistema.',
    '  No modificar ni eliminar la fila 1 (banners) ni la fila 2 (encabezados).',
    '  La fila 3 es un ejemplo — puede reemplazarla con datos reales.',
    '  Llenar desde la fila 3 en adelante. Maximo recomendado: 300 filas.',
]
for i, r in enumerate(reglas, start=10):
    wi.cell(row=i, column=1, value=r).font = Font(name='Arial', size=10)

wi['A23'] = 'PROCESO DE INSERCION'
wi['A23'].font = Font(name='Arial', bold=True, size=11, color='1e3a5f')
pasos = [
    '  1. Complete la hoja DATOS con los registros a insertar.',
    '  2. Guarde el archivo como .xlsx.',
    '  3. Entregue el archivo al administrador del sistema.',
    '  4. El administrador ejecutara el script de insercion en el servidor.',
    '  5. El script creara en orden: Lead -> Oportunidad -> Analisis -> Credito.',
    '  6. Se generara un reporte con los registros insertados y errores encontrados.',
]
for i, p in enumerate(pasos, start=24):
    wi.cell(row=i, column=1, value=p).font = Font(name='Arial', size=10)

# ─── HOJA 3: Valores_Validos ─────────────────────────────────────────────────
wv = wb.create_sheet('Valores_Validos')
wv.column_dimensions['A'].width = 28
wv.column_dimensions['B'].width = 70

for col in [1, 2]:
    labels = ['Campo', 'Valores Aceptados (escribir exactamente como aparece)']
    c = wv.cell(row=1, column=col, value=labels[col-1])
    c.font      = Font(name='Arial', bold=True, size=10, color='FFFFFF')
    c.fill      = fill(COLOR_CLIENTE)
    c.alignment = center
    c.border    = border
wv.row_dimensions[1].height = 22

alt = fill('EFF6FF')
valores = [
    ('genero',            'Masculino | Femenino | Otro'),
    ('estado_civil',      'Soltero | Casado | Divorciado | Viudo | Union Libre'),
    ('estado_puesto',     'Propietario | Interino | Contratado'),
    ('sector',            'Publico | Privado | Independiente'),
    ('nivel_academico',   'Primaria | Secundaria | Tecnico | Universidad | Postgrado'),
    ('source',            'Referido | Web | Llamada | Visita | Redes Sociales | Otro'),
    ('ana_divisa',        'CRC | USD'),
    ('ana_estado_pep',    'Pendiente | Aprobado | Rechazado | En Revision'),
    ('cred_tipo',         'Personal | Hipotecario | Prendario | Consumo | Empresarial'),
    ('cred_status',       'Activo | En Mora | Formalizado | Aprobado | Por firmar'),
    ('cred_garantia',     'Fiduciaria | Hipotecaria | Prendaria | Sin garantia'),
    ('opp_status',        'Pendiente | En Proceso | Ganada | Perdida | En Revision'),
    ('cred_poliza',       '1 (Con poliza) | 0 (Sin poliza)'),
    ('fechas',            'Formato: YYYY-MM-DD  (ej: 2026-04-01)'),
    ('montos',            'Solo numero sin comas ni simbolos  (ej: 3500000)'),
    ('tasa_anual',        'Porcentaje entero o decimal  (ej: 24 o 24.5)'),
]
for i, (campo, vals) in enumerate(valores, start=2):
    c1 = wv.cell(row=i, column=1, value=campo)
    c2 = wv.cell(row=i, column=2, value=vals)
    for c in [c1, c2]:
        c.font      = Font(name='Arial', size=10)
        c.border    = border
        c.alignment = left
        if i % 2 == 0:
            c.fill = alt
    wv.row_dimensions[i].height = 16

# ─── HOJA 4: Guia_Script ─────────────────────────────────────────────────────
wg = wb.create_sheet('Guia_Script')
wg.column_dimensions['A'].width = 90

wg['A1'] = 'Guia de Ejecucion del Script de Insercion'
wg['A1'].font = Font(name='Arial', bold=True, size=14, color='1e3a5f')

wg['A3'] = 'ORDEN DE INSERCION EN BASE DE DATOS'
wg['A3'].font = Font(name='Arial', bold=True, size=11, color='1e3a5f')
orden = [
    '  Paso 1: persons (lead)       -> tabla: persons, person_type_id = 1',
    '  Paso 2: opportunities        -> tabla: opportunities, vinculado por lead_cedula',
    '  Paso 3: analisis             -> tabla: analisis, vinculado por lead_id + opportunity_id',
    '  Paso 4: credits              -> tabla: credits, vinculado por lead_id + opportunity_id',
    '  Paso 5: plan_de_pagos        -> tabla: plan_de_pagos, generado automaticamente desde credito',
]
for i, o in enumerate(orden, start=4):
    wg.cell(row=i, column=1, value=o).font = Font(name='Arial', size=10)

wg['A10'] = 'NOTAS TECNICAS PARA EL SCRIPT'
wg['A10'].font = Font(name='Arial', bold=True, size=11, color='1e3a5f')
notas = [
    '  - opportunity.lead_cedula   = cedula de la columna CLIENTE',
    '  - analisis.lead_id          = persons.id obtenido por cedula',
    '  - analisis.opportunity_id   = opportunities.id recien insertada',
    '  - credit.lead_id            = persons.id obtenido por cedula',
    '  - credit.opportunity_id     = opportunities.id recien insertada',
    '  - credit.deductora_id       = buscar en deductoras WHERE nombre LIKE cred_deductora',
    '  - credit.tasa_anual         = cred_tasa_anual / 100  (convertir a decimal)',
    '  - El reference de oportunidad se genera automaticamente: YY-XXXXX-OP',
    '  - El reference de credito se genera automaticamente por el sistema',
    '  - plan_de_pago: calcular cuotas = monto * (tasa_mensual / (1-(1+tasa_mensual)^-plazo))',
]
for i, n in enumerate(notas, start=11):
    wg.cell(row=i, column=1, value=n).font = Font(name='Arial', size=10)

wb.save('public/plantilla_carga_masiva_completa.xlsx')
print('Generado: public/plantilla_carga_masiva_completa.xlsx')
